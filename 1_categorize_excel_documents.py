"""
Excel Document Categorizer & Analyzer
=====================================
Categorizes documents from Excel file metadata using Llama 3.3, then analyzes results.

Combines the functionality of:
- 1_categorize_documents.py (categorization)
- 2_analyze_categories.py (analysis)

Usage:
    python categorize_excel_documents.py input.xlsx
    python categorize_excel_documents.py input.xlsx -s "HR" -o output.xlsx
    python categorize_excel_documents.py --analyze-only results.xlsx
"""

import json
import re
import time
import argparse
from collections import Counter
from openpyxl import load_workbook

from config import DEFAULT_MODEL_ID
from oci_utils import (
    init_generative_ai_client,
    load_categories,
    create_chat_request,
    create_chat_details,
)


def predict_categories_batch(client, compartment_id, documents, categories):
    """
    Predict categories for a batch of documents using OCI GenAI LLM.
    """
    categories_list = "\n".join([f"- {cat}" for cat in categories])
    categories_enum = ", ".join([f'"{cat}"' for cat in categories] + ['"INVALID_CATEGORY"'])

    documents_text = ""
    for row_num, filename, doc_title in documents:
        filename_short = filename[:80] if len(filename) > 80 else filename
        doc_title_short = doc_title[:80] if len(doc_title) > 80 else doc_title
        documents_text += f'\n{row_num}: "{filename_short}" | "{doc_title_short}"'

    prompt = f"""
You are a strict document classification engine.

TASK:
For each document, select exactly ONE category from the allowed list below.
You MUST NOT create new categories, reword categories, or invent variants.
If no category fits, you MUST use "INVALID_CATEGORY".

ALLOWED_CATEGORIES (choose EXACTLY one of these):
{categories_list}

VALID OUTPUT VALUES:
{categories_enum}

SPECIAL RULES:
- If Filename contains "Copy.." you MUST use the category "Instruction to HR form"
- If the document is any kind of leaver, including "Leaver (unpaid)" or similar,
  you MUST output exactly "Leaver" and NOT create a subcategory.
- There is no "Change of Address" category, use "Change of Personal Details Form".
- Use "Addendum to Contract" for any likely changes of contract terms, like extensions, promotions, confirmations, etc. don't use "Contract"
- Use "Addendum to Contract" for paternity or maternity confirmations, RRA, Honorary scientist agreements
- If Filename starts with "Copy.." you MUST the category "Instruction to HR form"
- If PDR is in the title then it's "Personal Development Record
- Unsigned contracts go under "Employee Files - Starting Work"
- Anything with bank details go under "Bank Details"
- if the title is "Employee Unknown   AW..." or similar, then category is "HR Legacy At Work Archive DOB plus One Hundred Years"
- if however the title is like "Employee Unknown   SW ..." then category is "HR Legacy Starting Work Archive DOB plus One Hundred Years"
- if title or filenanme contains NSP or new Starter Form then classify as "Taleo New Starter Form"
- if the document has a title like 000676D8.pdf or similar then classify as "Compensation payments"

DOCUMENTS TO CLASSIFY
(format: row_number: "filename" | "title"):
{documents_text}

OUTPUT FORMAT (VERY IMPORTANT):
- Return ONLY a single JSON object.
- Keys: row numbers as strings (e.g. "2").
- Values: EXACT category names from ALLOWED_CATEGORIES above, or "INVALID_CATEGORY".
- No explanations, no comments, no extra keys, no text before or after the JSON.
"""

    chat_request = create_chat_request(prompt=prompt, max_tokens=4000, temperature=0.0)
    chat_detail = create_chat_details(
        chat_request, 
        model_id=DEFAULT_MODEL_ID, 
        compartment_id=compartment_id
    )

    try:
        chat_response = client.chat(chat_detail)
        response_text = (
            chat_response.data.chat_response.choices[0]
            .message.content[0]
            .text.strip()
        )

        json_match = re.search(r"\{.*\}", response_text, re.DOTALL)
        if json_match:
            response_text = json_match.group()

        predictions = json.loads(response_text)

        result = {}
        for row_str, category in predictions.items():
            try:
                row_num = int(row_str)
            except ValueError:
                continue

            if category == "INVALID_CATEGORY":
                result[row_num] = "Other"
            else:
                result[row_num] = category

        return result

    except json.JSONDecodeError as e:
        print(f"  ❌ Error parsing JSON response: {str(e)}")
        print(f"  Response was: {response_text[:200]}...")
        return {row_num: "Other" for row_num, _, _ in documents}

    except Exception as e:
        print(f"  ❌ Error predicting categories: {str(e)}")
        return {row_num: "Other" for row_num, _, _ in documents}


def categorize_documents(input_file, output_file, sheet_name="HR", 
                         filename_col=3, title_col=4, output_col=5,
                         batch_size=50):
    """
    Categorize documents in an Excel file.
    """
    print("=" * 80)
    print("Document Categorization")
    print("=" * 80)

    # Load categories
    print("\n📁 Loading categories...")
    categories = load_categories()
    print(f"✓ Loaded {len(categories)} categories")

    # Initialize OCI client
    print("\n🔌 Initializing OCI Generative AI client...")
    client, compartment_id = init_generative_ai_client()
    print("✓ Client initialized")

    # Load Excel file
    print(f"\n📊 Loading Excel file: {input_file}")
    workbook = load_workbook(input_file)
    sheet = workbook[sheet_name]
    print(f"✓ Loaded sheet: {sheet.title}")

    total_rows = sheet.max_row - 1
    print(f"✓ Found {total_rows} documents to categorize")

    # Collect documents
    print("\n📋 Collecting documents...")
    all_documents = []
    for row_num in range(2, sheet.max_row + 1):
        filename = sheet.cell(row=row_num, column=filename_col).value
        doc_title = sheet.cell(row=row_num, column=title_col).value

        if not filename and not doc_title:
            continue

        filename = str(filename) if filename else "N/A"
        doc_title = str(doc_title) if doc_title else "N/A"

        all_documents.append((row_num, filename, doc_title))

    print(f"✓ Collected {len(all_documents)} documents to process")

    # Process in batches
    total_batches = (len(all_documents) + batch_size - 1) // batch_size
    processed = 0

    print(f"\n🤖 Processing {total_batches} batches of up to {batch_size} documents...\n")

    for batch_num in range(0, len(all_documents), batch_size):
        batch = all_documents[batch_num : batch_num + batch_size]
        batch_index = batch_num // batch_size + 1

        print(
            f"Batch {batch_index}/{total_batches}: "
            f"Processing {len(batch)} documents (rows {batch[0][0]}-{batch[-1][0]})..."
        )

        predictions = predict_categories_batch(client, compartment_id, batch, categories)

        for row_num in predictions:
            category = predictions[row_num]
            sheet.cell(row=row_num, column=output_col).value = category
            processed += 1

        print(
            f"✓ Batch {batch_index} complete - {processed}/{len(all_documents)} documents categorized\n"
        )

        if batch_index < total_batches:
            time.sleep(1)

    # Save
    print(f"\n💾 Saving results to: {output_file}")
    workbook.save(output_file)
    print("✓ File saved successfully")

    return output_file, processed, total_batches


def analyze_results(excel_file, sheet_name="HR", original_col=2, predicted_col=5):
    """
    Analyze categorization results.
    """
    print("\n" + "=" * 80)
    print("ANALYSIS RESULTS")
    print("=" * 80)

    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]

    # Collect categories
    original_categories = []
    predicted_categories = []

    for row_num in range(2, sheet.max_row + 1):
        orig = sheet.cell(row=row_num, column=original_col).value
        pred = sheet.cell(row=row_num, column=predicted_col).value
        
        if orig:
            original_categories.append(str(orig))
        if pred:
            predicted_categories.append(str(pred))

    orig_counter = Counter(original_categories)
    pred_counter = Counter(predicted_categories)

    print("\n📊 ORIGINAL CATEGORIES (Column B) - Top 20:")
    print("-" * 50)
    for cat, count in orig_counter.most_common(20):
        print(f"  {cat}: {count}")

    print("\n📊 PREDICTED CATEGORIES (Column E) - Top 20:")
    print("-" * 50)
    for cat, count in pred_counter.most_common(20):
        print(f"  {cat}: {count}")

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total documents with original categories: {len(original_categories)}")
    print(f"Total documents with predicted categories: {len(predicted_categories)}")
    print(f"Unique original categories: {len(orig_counter)}")
    print(f"Unique predicted categories: {len(pred_counter)}")

    # Check for missing categories
    our_categories = set(load_categories())
    
    orig_not_in_ours = set(orig_counter.keys()) - our_categories
    pred_not_in_ours = set(pred_counter.keys()) - our_categories

    if orig_not_in_ours:
        print(f"\n⚠️  Original categories NOT in our {len(our_categories)}-category list: {len(orig_not_in_ours)}")
        print("Sample:")
        for cat in list(orig_not_in_ours)[:10]:
            print(f"  - {cat}")

    if pred_not_in_ours:
        print(f"\n⚠️  Predicted categories NOT in our list: {len(pred_not_in_ours)}")
        for cat in pred_not_in_ours:
            print(f"  - {cat}")

    # Category distribution
    print("\n📈 Category Distribution:")
    print(f"  Most common prediction: {pred_counter.most_common(1)[0] if pred_counter else 'N/A'}")
    print(f"  Categories with only 1 document: {sum(1 for c in pred_counter.values() if c == 1)}")

    return {
        "total_original": len(original_categories),
        "total_predicted": len(predicted_categories),
        "unique_original": len(orig_counter),
        "unique_predicted": len(pred_counter),
        "original_distribution": dict(orig_counter.most_common()),
        "predicted_distribution": dict(pred_counter.most_common()),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Categorize documents in Excel files using Llama 3.3"
    )
    parser.add_argument(
        "input_file",
        help="Path to input Excel file"
    )
    parser.add_argument(
        "-o", "--output",
        help="Output Excel file (default: input_categorized.xlsx)"
    )
    parser.add_argument(
        "-s", "--sheet",
        default="HR",
        help="Sheet name to process (default: HR)"
    )
    parser.add_argument(
        "--analyze-only",
        action="store_true",
        help="Only analyze existing results, don't categorize"
    )
    parser.add_argument(
        "--skip-analysis",
        action="store_true",
        help="Skip analysis after categorization"
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=50,
        help="Batch size for LLM processing (default: 50)"
    )
    
    args = parser.parse_args()
    
    # Determine output file
    if args.output:
        output_file = args.output
    else:
        import os
        base, ext = os.path.splitext(args.input_file)
        output_file = f"{base}_categorized{ext}"

    if args.analyze_only:
        # Just analyze existing file
        analyze_results(args.input_file, args.sheet)
    else:
        # Categorize and optionally analyze
        result_file, processed, batches = categorize_documents(
            args.input_file, 
            output_file,
            args.sheet,
            batch_size=args.batch_size
        )
        
        print("\n" + "=" * 80)
        print("CATEGORIZATION COMPLETE")
        print("=" * 80)
        print(f"Total documents processed: {processed}")
        print(f"Total API calls made: {batches} (batch processing)")
        print(f"Output file: {result_file}")

        if not args.skip_analysis:
            analyze_results(result_file, args.sheet)


if __name__ == "__main__":
    main()
